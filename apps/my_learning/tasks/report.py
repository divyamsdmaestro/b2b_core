import io
from datetime import datetime

import openpyxl
from django.core.files.storage import default_storage
from django.db.models import Q, Sum, Value
from django.db.models.functions import ExtractMonth, ExtractYear, TruncDate

from apps.common.tasks.base import BaseAppTask
from apps.event.config import TimePeriodChoices
from apps.learning.config import (
    BaseUploadStatusChoices,
    EvaluationTypeChoices,
    PlaygroundToolChoices,
    SubModuleTypeChoices,
)
from apps.meta.config import FeedBackTypeChoices
from apps.my_learning.config import AllBaseLearningTypeChoices, ApprovalTypeChoices, EnrollmentTypeChoices
from apps.tenant_service.middlewares import get_current_tenant_name

default_string = "-"


class ReportGenerationTask(BaseAppTask):
    """Task to generate report for specified user or user groups based on enrollment."""

    request_headers = None

    def setup_excel_header(self, sheet):
        """Function to setup the header for the excel sheet"""

        sheet.append(
            [
                "Employee Id",
                "Username",
                "Email-Id",
                "User Id",
                "Firstname",
                "Lastname",
                "Organization Name",
                "Business Unit",
                "Type",
                "Advanced Learning Path Name",
                "Learning Path Name",
                "Skill Traveller Name",
                "Playground Group Name",
                "Playground Name",
                "Skill Ontology Name",
                "Course Id",
                "Course Name",
                "Skills",
                "Proficiency",
                "Learning Points",
                "Duration",
                "Lab Duration",
                "Enrolment Date",
                "Started Date",
                "Completion Date",
                "Deadline",
                "Status",
                "Enrollment Method",
                "Video Progress",
                "Assessment Progress",
                "Lab Progress",
                "Scorm Progress",
                "Assessment Result",
                "Assessment Score",
                "Availed Attempts",
                "Assignment Progress",
                "Assignment Result",
                "Assignment Score",
            ]
        )

    def process_assignment_tracker(self, user, tool, name, filter_params, assignment_trackers):
        """Function to process the assignment tracker of core & ccms instances."""

        if tracker := (
            user.related_assignment_trackers.filter(is_completed=True, **filter_params).order_by("-progress").first()
        ):
            if tool == PlaygroundToolChoices.yaksha:
                if yaksha_schedule := tracker.related_assignment_yaksha_schedules.first():
                    if assessment_result := yaksha_schedule.related_assignment_yaksha_results.order_by(
                        "-progress"
                    ).first():
                        setattr(tracker, "score", assessment_result.progress)
            elif tool == PlaygroundToolChoices.mml:
                if submission := tracker.related_assignment_submissions.order_by("-progress").first():
                    setattr(tracker, "score", submission.progress)
            setattr(tracker, "assignment_name", name)
            assignment_trackers.append(tracker)

    def get_ccms_trackers(self, learning_type, ccms_id, user, learning_data):
        """Get learning trackers, assignment, assessment, and assessment result."""

        from apps.my_learning.helpers import RELATED_TRACKER_NAMES
        from apps.my_learning.models import CourseAssessmentTracker, CourseSubModuleTracker

        user_learning_tracker, assignment_trackers, assessment_data_list = None, [], []
        file_submissions, sub_module_details = [], {}
        try:
            user_learning_tracker = getattr(user, RELATED_TRACKER_NAMES[learning_type]).filter(ccms_id=ccms_id).first()
            match learning_type:
                case EnrollmentTypeChoices.course:
                    assessment_ids, assessment_details = [], {}
                    for data in learning_data["assessment"]:
                        assessment_ids.append(data["uuid"])
                        assessment_details[data["uuid"]] = data["name"]
                    course_trackers = user.related_user_course_trackers.filter(ccms_id=ccms_id)
                    course_assessment_trackers = CourseAssessmentTracker.objects.filter(
                        Q(course_tracker__in=course_trackers) | Q(module_tracker__course_tracker__in=course_trackers),
                        ccms_id__in=assessment_ids,
                    )
                    for assessment_tracker in course_assessment_trackers:
                        if yaksha_schedules := assessment_tracker.related_ca_yaksha_schedules.first():
                            if assessment_result := yaksha_schedules.related_ca_yaksha_results.order_by(
                                "-progress"
                            ).first():
                                assessment_tracker.assessment_name = assessment_details.get(
                                    str(assessment_tracker.ccms_id), default_string
                                )
                                assessment_data_list.append([assessment_tracker, assessment_result])
                    if user_learning_tracker:
                        sub_module_ids = []
                        for data in learning_data["sub_module"]:
                            if data["type"]["id"] == "file_submission":
                                sub_module_uuid = data["uuid"]
                                sub_module_ids.append(sub_module_uuid)
                                sub_module_details[sub_module_uuid] = {
                                    "name": data["name"],
                                    "evaluation_type": data["evaluation_type"],
                                }
                        sub_module_trackers = CourseSubModuleTracker.objects.filter(
                            ccms_id__in=sub_module_ids,
                            module_tracker__course_tracker=user_learning_tracker,
                        )
                        for tracker in sub_module_trackers:
                            if submission := tracker.related_sub_module_submissions.order_by("-progress").first():
                                file_submissions.append(submission)
                case EnrollmentTypeChoices.learning_path:
                    assessment_ids, assessment_details = [], {}
                    for data in learning_data["assessment"]:
                        assessment_ids.append(data["uuid"])
                        assessment_details[data["uuid"]] = data["name"]
                    lp_assessment_trackers = user.related_lp_assessment_trackers.filter(ccms_id__in=assessment_ids)
                    for assessment_tracker in lp_assessment_trackers:
                        if yaksha_schedules := assessment_tracker.related_lpa_yaksha_schedules.first():
                            if assessment_result := yaksha_schedules.related_lpa_yaksha_results.order_by(
                                "-progress"
                            ).first():
                                assessment_tracker.assessment_name = assessment_details.get(
                                    str(assessment_tracker.ccms_id), default_string
                                )
                                assessment_data_list.append([assessment_tracker, assessment_result])
            for learning_assignment in learning_data.get("assignment", []):
                uuid = learning_assignment["assignment"]["uuid"]
                tool = learning_assignment["assignment"]["tool"]
                name = learning_assignment["assignment"]["name"]
                self.process_assignment_tracker(user, tool, name, {"ccms_id": uuid}, assignment_trackers)
        except Exception as e:
            self.logger.error(e)
        return user_learning_tracker, assignment_trackers, assessment_data_list, file_submissions, sub_module_details

    def get_core_trackers(self, learning_obj, learning_obj_type, user):
        """Get learning trackers, assignment, assessment, and assessment result."""

        from apps.learning.models import (
            CourseAssessment,
            CourseAssignment,
            LPAssessment,
            LPAssignment,
            STAssessment,
            STAssignment,
        )
        from apps.my_learning.helpers import RELATED_TRACKER_NAMES
        from apps.my_learning.models import CourseAssessmentTracker, CourseSubModuleTracker

        user_learning_tracker, assignment_trackers, assessment_data, file_submissions = None, [], [], []
        try:
            user_learning_tracker = (
                getattr(user, RELATED_TRACKER_NAMES[learning_obj_type])
                .filter(**{learning_obj_type: learning_obj})
                .first()
            )
            learning_assignments = []
            match learning_obj_type:
                case EnrollmentTypeChoices.course:
                    learning_assessments = CourseAssessment.objects.filter(
                        Q(course=learning_obj) | Q(module__course=learning_obj)
                    )
                    learning_assignments = CourseAssignment.objects.filter(
                        Q(course=learning_obj) | Q(module__course=learning_obj)
                    )
                    course_trackers = user.related_user_course_trackers.filter(course=learning_obj)
                    course_assessment_trackers = CourseAssessmentTracker.objects.filter(
                        Q(course_tracker__in=course_trackers) | Q(module_tracker__course_tracker__in=course_trackers),
                        assessment__in=learning_assessments,
                    )
                    for assessment_tracker in course_assessment_trackers:
                        if yaksha_schedules := assessment_tracker.related_ca_yaksha_schedules.first():
                            if assessment_result := yaksha_schedules.related_ca_yaksha_results.order_by(
                                "-progress"
                            ).first():
                                assessment_data.append([assessment_tracker, assessment_result])
                    if user_learning_tracker:
                        sub_module_trackers = CourseSubModuleTracker.objects.filter(
                            sub_module__type=SubModuleTypeChoices.file_submission,
                            module_tracker__course_tracker=user_learning_tracker,
                        )
                        for tracker in sub_module_trackers:
                            if submission := tracker.related_sub_module_submissions.order_by("-progress").first():
                                file_submissions.append(submission)
                case EnrollmentTypeChoices.learning_path:
                    learning_assessments = LPAssessment.objects.filter(
                        Q(learning_path=learning_obj) | Q(lp_course__learning_path=learning_obj)
                    )
                    learning_assignments = LPAssignment.objects.filter(
                        Q(learning_path=learning_obj) | Q(lp_course__learning_path=learning_obj)
                    )
                    lp_assessment_trackers = user.related_lp_assessment_trackers.filter(
                        assessment__in=learning_assessments
                    )
                    for assessment_tracker in lp_assessment_trackers:
                        if yaksha_schedules := assessment_tracker.related_lpa_yaksha_schedules.first():
                            if assessment_result := yaksha_schedules.related_lpa_yaksha_results.order_by(
                                "-progress"
                            ).first():
                                assessment_data.append([assessment_tracker, assessment_result])
                case EnrollmentTypeChoices.skill_traveller:
                    learning_assessments = STAssessment.objects.filter(
                        Q(skill_traveller=learning_obj) | Q(st_course__skill_traveller=learning_obj)
                    )
                    learning_assignments = STAssignment.objects.filter(
                        Q(skill_traveller=learning_obj) | Q(st_course__skill_traveller=learning_obj)
                    )
                    st_assessment_trackers = user.related_st_assessment_trackers.filter(
                        assessment__in=learning_assessments
                    )
                    for assessment_tracker in st_assessment_trackers:
                        if yaksha_schedules := assessment_tracker.related_sta_yaksha_schedules.first():
                            if assessment_result := yaksha_schedules.related_sta_yaksha_results.order_by(
                                "-progress"
                            ).first():
                                assessment_data.append([assessment_tracker, assessment_result])
                case EnrollmentTypeChoices.assignment:
                    if user_learning_tracker:
                        if learning_obj.tool == PlaygroundToolChoices.yaksha:
                            if yaksha_schedule := user_learning_tracker.related_assignment_yaksha_schedules.first():
                                if assessment_result := yaksha_schedule.related_assignment_yaksha_results.order_by(
                                    "-progress"
                                ).first():
                                    assessment_data.append([user_learning_tracker, assessment_result])
                        elif learning_obj.tool == PlaygroundToolChoices.mml:
                            if submission := user_learning_tracker.related_assignment_submissions.order_by(
                                "-progress"
                            ).first():
                                file_submissions.append(submission)
            for learning_assignment in learning_assignments:
                assignment_obj = learning_assignment.assignment
                self.process_assignment_tracker(
                    user,
                    assignment_obj.tool,
                    assignment_obj.name,
                    {"assignment_id": assignment_obj.id},
                    assignment_trackers,
                )
        except Exception as e:
            self.logger.error(e)
        return user_learning_tracker, assignment_trackers, assessment_data, file_submissions

    def get_approval_type(self, enrollment_instance):
        """Returns the enrollment approval type"""

        if enrollment_instance.approval_type in [
            ApprovalTypeChoices.self_enrolled,
            ApprovalTypeChoices.tenant_admin,
            ApprovalTypeChoices.super_admin,
        ]:
            return ApprovalTypeChoices.get_choice(enrollment_instance.approval_type).label
        else:
            return default_string

    def get_user_tracker_data(self, user_learning_tracker, learning_points):
        """Returns the progress, started_date, completion_date, learning_points, status from user tracker."""

        if user_learning_tracker:
            video_progress = user_learning_tracker.progress
            started_date = (
                user_learning_tracker.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if user_learning_tracker.created_at
                else default_string
            )
            completion_date = (
                user_learning_tracker.completion_date.strftime("%Y-%m-%d %H:%M:%S")
                if user_learning_tracker.completion_date
                else default_string
            )
            learning_points, status = (
                (learning_points, "Completed") if user_learning_tracker.is_completed else (default_string, "Started")
            )
        else:
            video_progress = started_date = completion_date = learning_points = default_string
            status = "Not Started"
        return video_progress, started_date, completion_date, learning_points, status

    def update_assessment_data_list(self, tracker, result_obj, name, availed_attempts, progress, results, scores):
        """Function to update lists of assessment progress, scores and results."""

        if tracker.allowed_attempt is not None and tracker.available_attempt is not None:
            attempt = tracker.allowed_attempt + getattr(tracker, "reattempt_count", 0) - tracker.available_attempt
            availed_attempts.append(f"{name} - {attempt}")
        if result_obj.is_pass:
            progress.append(f"{name} - 100")
            results.append(f"{name} - Passed")
        else:
            progress.append(f"{name} - 100")
            results.append(f"{name} - Failed")
        scores.append(f"{name} - {result_obj.progress}")

    def update_assignment_data_list(self, tracker, assign_type, name, results, progress, scores):
        """Function to update lists of assignment progress, scores and results."""

        if assign_type == EnrollmentTypeChoices.assignment:
            submission_status, submission_progress = ("Passed", 100) if tracker.is_pass else ("Failed", 0)
            submission_score = getattr(tracker, "score", "N/A")
        elif assign_type == EvaluationTypeChoices.evaluated:
            if tracker.is_reviewed:
                submission_status, submission_progress = ("Passed", 100) if tracker.is_pass else ("Failed", 0)
                submission_score = tracker.progress
            else:
                submission_status = "Evaluation Pending"
                submission_progress = submission_score = "N/A"
        else:
            submission_status, submission_progress, submission_score = "Passed", 100, "N/A"
        results.append(f"{name} - {submission_status}")
        progress.append(f"{name} - {submission_progress}")
        scores.append(f"{name} - {submission_score}")

    def populate_excel_row(
        self,
        learning_type,
        sheet,
        enrollment_instance,
        user,
        learning_obj=None,
        learning_data=None,
        is_ccms_obj=False,
        ccms_id=None,
    ):
        """Function to setup the values in excel file."""

        from apps.access.models import UserDetail
        from apps.my_learning.helpers import convert_sec_to_hms

        course_name = course_code = lp_name = alp_name = st_name = so_name = ag_name = assignment_name = default_string
        if not is_ccms_obj:
            if learning_type == EnrollmentTypeChoices.skill_ontology:
                cur_skill = learning_obj.current_skill_detail
                desired_skill = learning_obj.desired_skill_detail
                skill_list = f"Current Skill: {cur_skill.skill.name}, Desired Skill: {desired_skill.skill.name}"
                proficiency = f"Current Skill: {cur_skill.proficiency}, Desired Skill: {desired_skill.proficiency}"
                so_name = enrollment_instance.skill_ontology.name
            else:
                proficiency = learning_obj.proficiency
                skill_list = ", ".join(learning_obj.skill.all().values_list("name", flat=True))
            (user_learning_tracker, assignment_trackers, assessment_data, file_submissions) = self.get_core_trackers(
                learning_obj, learning_type, user
            )
            learning_points = getattr(learning_obj, "learning_points", default_string)
            duration = getattr(learning_obj, "duration", None)
        else:
            artifact_details = learning_data[learning_type]
            proficiency, duration = artifact_details["proficiency"], artifact_details["duration"]
            learning_points, skill_list = artifact_details["learning_points"], ", ".join(artifact_details["skill"])
            (
                user_learning_tracker,
                assignment_trackers,
                assessment_data,
                file_submissions,
                sub_module_details,
            ) = self.get_ccms_trackers(learning_type, ccms_id, user, learning_data)
        match learning_type:
            case EnrollmentTypeChoices.course:
                if is_ccms_obj:
                    course_name, course_code = artifact_details["name"], artifact_details["code"]
                    lp_name, alp_name = artifact_details["lp_name"], artifact_details["alp_name"]
                else:
                    course_name, course_code = learning_obj.name, learning_obj.code
                    lp_name = getattr(learning_obj, "lp_name", default_string)
                    alp_name = getattr(learning_obj, "alp_name", default_string)
                    st_name = getattr(learning_obj, "st_name", default_string)
            case EnrollmentTypeChoices.learning_path:
                if not is_ccms_obj:
                    lp_name, alp_name = learning_obj.name, getattr(learning_obj, "alp_name", default_string)
                else:
                    lp_name, alp_name = artifact_details["name"], artifact_details["alp_name"]
            case EnrollmentTypeChoices.advanced_learning_path:
                alp_name = artifact_details["name"] if is_ccms_obj else learning_obj.name
            case EnrollmentTypeChoices.skill_traveller:
                st_name = learning_obj.name
            case EnrollmentTypeChoices.assignment_group:
                ag_name, video_progress = learning_obj.name, default_string
            case EnrollmentTypeChoices.assignment:
                assignment_name, ag_name = learning_obj.name, getattr(learning_obj, "ag_name", default_string)
                video_progress = default_string
        availed_attempts, assess_progress, assess_results, assess_scores = [], [], [], []
        assign_results, assign_scores, assign_progress = [], [], []
        try:
            for assessment in assessment_data:
                tracker, result = assessment
                if is_ccms_obj:
                    assess_name = tracker.assessment_name
                elif learning_type == EnrollmentTypeChoices.assignment:
                    assess_name = assignment_name
                else:
                    assess_name = tracker.assessment.name
                self.update_assessment_data_list(
                    tracker,
                    result,
                    assess_name,
                    availed_attempts,
                    assess_progress,
                    assess_results,
                    assess_scores,
                )
            for assignment_tracker in assignment_trackers:
                self.update_assignment_data_list(
                    assignment_tracker,
                    EnrollmentTypeChoices.assignment,
                    getattr(assignment_tracker, "assignment_name", "N/A"),
                    assign_results,
                    assign_progress,
                    assign_scores,
                )
            for submission in file_submissions:
                if is_ccms_obj:
                    sub_module_uuid = submission.sub_module_tracker.ccms_id
                    evaluation_type = sub_module_details[str(sub_module_uuid)]["evaluation_type"]
                    name = sub_module_details[str(sub_module_uuid)]["name"]
                else:
                    if learning_type == EnrollmentTypeChoices.assignment:
                        related_instance = submission.assignment_tracker.assignment
                    else:
                        related_instance = submission.sub_module_tracker.sub_module
                    evaluation_type, name = related_instance.evaluation_type, related_instance.name
                self.update_assignment_data_list(
                    submission,
                    evaluation_type,
                    name,
                    assign_results,
                    assign_progress,
                    assign_scores,
                )
        except Exception as e:
            self.logger.error(e)
            availed_attempts = assess_progress = assess_results = assess_scores = [default_string]
            assign_results = assign_scores = assign_progress = [default_string]
        video_progress, started_date, completion_date, learning_points, status = self.get_user_tracker_data(
            user_learning_tracker, learning_points
        )
        employee_id = user_id = business_unit_name = default_string
        if user_detail_obj := UserDetail.objects.filter(user=user).first():
            employee_id, user_id = user_detail_obj.employee_id, user_detail_obj.user_id_number
            business_unit_name = user_detail_obj.business_unit_name
        sheet.append(
            [
                employee_id,
                user.username,
                user.email,
                user_id,
                user.first_name,
                user.last_name,
                get_current_tenant_name(),
                business_unit_name,
                EnrollmentTypeChoices.get_choice(enrollment_instance.learning_type).label,
                alp_name,
                lp_name,
                st_name,
                ag_name,
                assignment_name,
                so_name,
                course_code,
                course_name,
                skill_list if skill_list else default_string,
                proficiency,
                learning_points,
                convert_sec_to_hms(duration) if duration else default_string,
                default_string,
                str(enrollment_instance.action_date) if enrollment_instance.action_date else default_string,
                started_date,
                completion_date,
                enrollment_instance.end_date.strftime("%Y-%m-%d %H:%M:%S")
                if enrollment_instance.end_date
                else default_string,
                status,
                self.get_approval_type(enrollment_instance),
                video_progress,
                ", ".join(assess_progress),
                default_string,
                default_string,
                ", ".join(assess_results),
                ", ".join(assess_scores),
                ", ".join(availed_attempts),
                ", ".join(assign_progress),
                ", ".join(assign_results),
                ", ".join(assign_scores),
            ]
        )

    def populate_excel_row_lp(self, learning_obj, **kwargs):
        """Function to populate excel row for lp and related learnings."""

        from apps.learning.models import Course

        self.populate_excel_row(learning_obj=learning_obj, **kwargs)
        alp_name = getattr(learning_obj, "alp_name", default_string)
        courses = Course.objects.filter(related_learning_path_courses__learning_path=learning_obj).annotate(
            lp_name=Value(learning_obj.name), alp_name=Value(alp_name)
        )
        kwargs.update({"learning_type": EnrollmentTypeChoices.course})
        for course in courses:
            self.populate_excel_row(learning_obj=course, **kwargs)

    def populate_excel_row_alp(self, learning_obj, **kwargs):
        """Function to populate excel row for alp and related learnings."""

        from apps.learning.models import LearningPath

        self.populate_excel_row(learning_obj=learning_obj, **kwargs)
        lps = LearningPath.objects.filter(related_alp_learning_paths__advanced_learning_path=learning_obj).annotate(
            alp_name=Value(learning_obj.name)
        )
        kwargs.update({"learning_type": EnrollmentTypeChoices.learning_path})
        for lp in lps:
            self.populate_excel_row_lp(learning_obj=lp, **kwargs)

    def populate_excel_row_st(self, learning_obj, **kwargs):
        """Function to populate excel row for st and related learnings."""

        from apps.learning.models import Course

        self.populate_excel_row(learning_obj=learning_obj, **kwargs)
        courses = Course.objects.filter(related_skill_traveller_courses__skill_traveller=learning_obj).annotate(
            st_name=Value(learning_obj.name)
        )
        kwargs.update({"learning_type": EnrollmentTypeChoices.course})
        for course in courses:
            self.populate_excel_row(learning_obj=course, **kwargs)

    def populate_excel_row_ag(self, learning_obj, **kwargs):
        """Function to populate excel row for assignment group and related learnings."""

        from apps.learning.models import Assignment

        self.populate_excel_row(learning_obj=learning_obj, **kwargs)
        assignments = Assignment.objects.filter(related_assignment_relations__assignment_group=learning_obj).annotate(
            ag_name=Value(learning_obj.name)
        )
        kwargs.update({"learning_type": EnrollmentTypeChoices.assignment})
        for assignment in assignments:
            self.populate_excel_row(learning_obj=assignment, **kwargs)

    def populate_excel_row_so(self, learning_obj, **kwargs):
        """Function to populate excel row for skill ontology and related learnings."""

        self.populate_excel_row(learning_obj=learning_obj, **kwargs)
        kwargs.pop("learning_type")
        for course in learning_obj.course.all():
            self.populate_excel_row(learning_obj=course, learning_type=EnrollmentTypeChoices.course, **kwargs)
        for lp in learning_obj.learning_path.all():
            self.populate_excel_row_lp(learning_obj=lp, learning_type=EnrollmentTypeChoices.learning_path, **kwargs)
        for alp in learning_obj.advanced_learning_path.all():
            self.populate_excel_row_alp(
                learning_obj=alp, learning_type=EnrollmentTypeChoices.advanced_learning_path, **kwargs
            )
        for skill_traveller in learning_obj.skill_traveller.all():
            self.populate_excel_row_st(
                learning_obj=skill_traveller, learning_type=EnrollmentTypeChoices.skill_traveller, **kwargs
            )
        for assignment in learning_obj.assignment.all():
            self.populate_excel_row(learning_obj=assignment, learning_type=EnrollmentTypeChoices.assignment, **kwargs)
        for assignment_group in learning_obj.assignment_group.all():
            self.populate_excel_row_ag(
                learning_obj=assignment_group, learning_type=EnrollmentTypeChoices.assignment_group, **kwargs
            )

    def populate_excel_row_lp_ccms(self, learning_data, **kwargs):
        """Function to populate excel row for ccms lp and related learnings."""

        self.populate_excel_row(learning_data=learning_data, **kwargs)
        kwargs.update({"learning_type": EnrollmentTypeChoices.course})
        for course in learning_data["courses"]:
            kwargs["ccms_id"] = course["course"]["uuid"]
            self.populate_excel_row(learning_data=course, **kwargs)

    def populate_excel_row_alp_ccms(self, learning_data, **kwargs):
        """Function to populate excel row for ccms alp and related learnings."""

        self.populate_excel_row(learning_data=learning_data, **kwargs)
        kwargs.update({"learning_type": EnrollmentTypeChoices.learning_path})
        for learning_path in learning_data["learning_paths"]:
            kwargs["ccms_id"] = learning_path["learning_path"]["uuid"]
            self.populate_excel_row_lp_ccms(learning_data=learning_path, **kwargs)

    def trigger_function_based_on_learning_type(self, **kwargs):
        """Function to trigger populate_excel_row func. Just a DRY Stuff."""

        core_function_mapping = {
            EnrollmentTypeChoices.course: self.populate_excel_row,
            EnrollmentTypeChoices.learning_path: self.populate_excel_row_lp,
            EnrollmentTypeChoices.advanced_learning_path: self.populate_excel_row_alp,
            EnrollmentTypeChoices.skill_traveller: self.populate_excel_row_st,
            EnrollmentTypeChoices.assignment: self.populate_excel_row,
            EnrollmentTypeChoices.assignment_group: self.populate_excel_row_ag,
            EnrollmentTypeChoices.skill_ontology: self.populate_excel_row_so,
        }
        ccms_function_mapping = {
            EnrollmentTypeChoices.course: self.populate_excel_row,
            EnrollmentTypeChoices.learning_path: self.populate_excel_row_lp_ccms,
            EnrollmentTypeChoices.advanced_learning_path: self.populate_excel_row_alp_ccms,
        }
        enrollment = kwargs["enrollment_instance"]
        learning_type = enrollment.learning_type
        if learning_type in ccms_function_mapping and enrollment.is_ccms_obj:
            ccms_function_mapping[learning_type](**kwargs)
        elif learning_type in core_function_mapping and kwargs.get("learning_obj"):
            core_function_mapping[learning_type](**kwargs)

    def populate_excel_sheet(self, sheet, enrolled_objs, user_ids, is_user_report):
        """Function to populate the report sheet with the given courses."""

        from apps.learning.helpers import get_ccms_retrieve_details

        for enrollment in enrolled_objs:
            if enrollment.user:
                users = [enrollment.user]
            elif enrollment.user_group:
                if not is_user_report:
                    users = enrollment.user_group.members.all()
                else:
                    users = enrollment.user_group.members.filter(id__in=user_ids)
            else:
                continue
            kwargs = {
                "sheet": sheet,
                "enrollment_instance": enrollment,
                "learning_type": enrollment.learning_type,
            }
            if enrollment.is_ccms_obj:
                success, learning_data = get_ccms_retrieve_details(
                    learning_type=f"core_{enrollment.learning_type}",
                    instance_id=enrollment.ccms_id,
                    request=self.request_headers,
                )
                if not success:
                    self.logger.error(f"Report Generation Task failed: {learning_data}")
                    continue
                kwargs.update(
                    {
                        "learning_data": learning_data["data"],
                        "is_ccms_obj": True,
                        "ccms_id": enrollment.ccms_id,
                    }
                )
            else:
                kwargs.update({"learning_obj": getattr(enrollment, enrollment.learning_type)})
            for user in users:
                kwargs.update({"user": user})
                self.trigger_function_based_on_learning_type(**kwargs)

    def run(self, report_instance_id, db_name, request_headers, **kwargs):
        """Run handler."""

        from apps.my_learning.models import Enrollment, Report

        self.switch_db(db_name)
        self.logger.info(f"Executing ReportGenerationTask on {db_name} database.")

        self.request_headers = request_headers
        report_instance = Report.objects.get(id=report_instance_id)
        report_instance.status = BaseUploadStatusChoices.in_progress
        report_instance.save()
        enrolled_objs = Enrollment.objects.filter(is_enrolled=True)
        if not report_instance.data["is_entire_learnings"]:
            enrolled_objs = enrolled_objs.filter(
                Q(course_id__in=report_instance.data["course"])
                | Q(learning_path_id__in=report_instance.data["learning_path"])
                | Q(advanced_learning_path_id__in=report_instance.data["advanced_learning_path"])
                | Q(skill_traveller_id__in=report_instance.data["skill_traveller"])
                | Q(assignment_id__in=report_instance.data["assignment"])
                | Q(assignment_group_id__in=report_instance.data["assignment_group"])
            )
        if not report_instance.data["is_date_skipped"]:
            enrolled_objs = enrolled_objs.filter(
                action_date__gte=report_instance.start_date, action_date__lte=report_instance.end_date
            )
        is_user_report = False
        if report_instance.data["user"]:
            enrolled_objs = enrolled_objs.filter(
                Q(user__in=report_instance.data["user"]) | Q(user_group__members__in=report_instance.data["user"])
            ).distinct()
            is_user_report = True
        elif report_instance.data["user_group"]:
            enrolled_objs = enrolled_objs.filter(user_group_id__in=report_instance.data["user_group"])
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self.setup_excel_header(sheet)
            self.populate_excel_sheet(sheet, enrolled_objs, report_instance.data["user"], is_user_report)
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            uploaded_file = default_storage.save(f"files/{db_name}/report/{report_instance.name}.xlsx", excel_file)
            report_instance.file_url = default_storage.url(uploaded_file)
            report_instance.status = BaseUploadStatusChoices.completed
            report_instance.save()
            report_instance.call_report_emailtask(file_path=uploaded_file, db_name=db_name)
        except Exception as e:
            self.logger.error(e)
            report_instance.status = BaseUploadStatusChoices.failed
            report_instance.save()
        return True


class FileSubmissionReportGenerationTask(BaseAppTask):
    """Task to generate report for specified user or user groups based on File submissions."""

    def setup_excel_header(self, sheet, report_instance):
        """Function to setup the header for the excel sheet"""

        row_data = [
            "Username",
            "User Group",
            "Course Name",
            "Module Name",
            "BlobUrl",
            "Submission Date",
        ]
        if report_instance.data["is_evaluated"]:
            row_data.extend(
                [
                    "PassThreshold",
                    "Availed Attempts",
                    "Score Points",
                    "Evaluated Date",
                    "Evaluation Status",
                ]
            )
        else:
            row_data.extend(["Evaluation Status"])
        sheet.append(row_data)

    def populate_excel_sheet(self, sheet, report_instance, sub_module_trackers):
        """Function to setup the values in excel file."""

        for tracker in sub_module_trackers:
            user = tracker.module_tracker.course_tracker.user
            user_group = user.related_user_groups.first()
            for submission in tracker.related_sub_module_submissions.order_by("attempt"):
                for file in submission.files.all():
                    row_data = [
                        user.username,
                        user_group.name if user_group else default_string,
                        tracker.sub_module.module.course.name,
                        tracker.sub_module.module.name,
                        file.file_url,
                        submission.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    ]
                    if report_instance.data["is_evaluated"]:
                        row_data.extend(
                            [
                                submission.pass_percentage,
                                submission.attempt,
                                submission.progress,
                                submission.modified_at.strftime("%Y-%m-%d %H:%M:%S")
                                if submission.is_reviewed
                                else default_string,
                                ("Passed" if submission.is_pass else "Failed")
                                if submission.is_reviewed
                                else "Pending",
                            ]
                        )
                    else:
                        row_data.extend(["Passed" if tracker.is_completed else "Failed"])
                    sheet.append(row_data)

    def run(self, report_instance_id, db_name, **kwargs):
        """Run handler."""

        from apps.my_learning.models import CourseSubModuleTracker, Report

        self.switch_db(db_name)
        self.logger.info(f"Executing FileSubmissionReportGenerationTask on {db_name} database.")

        report_instance = Report.objects.get(id=report_instance_id)
        sub_module_trackers = CourseSubModuleTracker.objects.filter(
            sub_module__type=SubModuleTypeChoices.file_submission
        )
        if report_instance.data["is_evaluated"]:
            sub_module_trackers = sub_module_trackers.filter(
                sub_module__evaluation_type=EvaluationTypeChoices.evaluated
            )
        else:
            sub_module_trackers = sub_module_trackers.filter(
                sub_module__evaluation_type=EvaluationTypeChoices.non_evaluated
            )
        if not report_instance.data["is_entire_learnings"]:
            sub_module_trackers = sub_module_trackers.filter(
                Q(sub_module__module__course_id__in=report_instance.data["course"])
                | Q(
                    sub_module__module__course__related_learning_path_courses__learning_path_id__in=report_instance.data[  # noqa
                        "learning_path"
                    ]
                )
                | Q(
                    sub_module__module__course__related_learning_path_courses__learning_path__related_alp_learning_paths__advanced_learning_path_id__in=report_instance.data[  # noqa
                        "advanced_learning_path"
                    ]
                )
            )
        if not report_instance.data["is_date_skipped"]:
            sub_module_trackers = sub_module_trackers.filter(
                related_sub_module_submissions__created_at__date__gte=report_instance.start_date,
                related_sub_module_submissions__created_at__date__lte=report_instance.end_date,
            ).distinct()
        if report_instance.data["user"] or report_instance.data["user_group"]:
            sub_module_trackers = sub_module_trackers.filter(
                Q(module_tracker__course_tracker__user_id__in=report_instance.data["user"])
                | Q(
                    module_tracker__course_tracker__user__related_user_groups__id__in=report_instance.data[
                        "user_group"
                    ]
                )
            )
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self.setup_excel_header(sheet, report_instance)
            self.populate_excel_sheet(sheet, report_instance, sub_module_trackers)
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            uploaded_file = default_storage.save(f"files/{db_name}/report/{report_instance.name}.xlsx", excel_file)
            report_instance.file_url = default_storage.url(uploaded_file)
            report_instance.call_report_emailtask(file_path=uploaded_file, db_name=db_name)
            report_instance.status = BaseUploadStatusChoices.completed
            report_instance.save()
        except Exception:
            report_instance.status = BaseUploadStatusChoices.failed
            report_instance.save()
        return True


class FeedbackReportGenerationTask(BaseAppTask):
    """Celery task to Generate Feedback Report"""

    @staticmethod
    def setup_excel_header(sheet):
        """Function to Populate Headers"""

        sheet.append(
            [
                "Username",
                "Firstname",
                "Lastname",
                "Course",
                "Learning Path",
                "Advanced Learning Path",
                "Feedback Template",
                "Question",
                "Response",
            ]
        )

    @staticmethod
    def populate_excel_sheet(sheet, feedback_objs):
        """Function to Populate report data in Excel"""

        from apps.learning.helpers import LEARNING_INSTANCE_MODELS

        for feedback in feedback_objs:
            course_name = lp_name = alp_name = default_string

            obj = LEARNING_INSTANCE_MODELS[feedback.learning_type].objects.filter(id=feedback.learning_type_id).first()

            match feedback.learning_type:
                case AllBaseLearningTypeChoices.course:
                    course_name = obj.name
                case AllBaseLearningTypeChoices.learning_path:
                    lp_name = obj.name
                case AllBaseLearningTypeChoices.advanced_learning_path:
                    alp_name = obj.name

            if feedback.question.response_type == FeedBackTypeChoices.choice:
                response = feedback.choice.choice
            else:
                response = feedback.text

            sheet.append(
                [
                    feedback.user.username,
                    feedback.user.first_name,
                    feedback.user.last_name,
                    course_name,
                    lp_name,
                    alp_name,
                    feedback.question.feedback_template.name,
                    feedback.question.question,
                    response,
                ]
            )

    def run(self, db_name, report_instance_id):
        """Run Handler"""

        from apps.my_learning.models import FeedbackResponse, Report

        self.switch_db(db_name)
        self.logger.info(f"Executing ReportGenerationTask on {db_name} database.")

        report_instance = Report.objects.get(id=report_instance_id)
        # TODO: Need to provide ccms support.
        feedback_objs = FeedbackResponse.objects.filter(is_ccms_obj=False)

        if not report_instance.data["is_entire_learnings"]:
            feedback_objs = feedback_objs.filter(
                Q(learning_type_id__in=report_instance.data["course"], learning_type=AllBaseLearningTypeChoices.course)
                | Q(
                    learning_type_id__in=report_instance.data["learning_path"],
                    learning_type=AllBaseLearningTypeChoices.learning_path,
                )
                | Q(
                    learning_type_id__in=report_instance.data["advanced_learning_path"],
                    learning_type=AllBaseLearningTypeChoices.advanced_learning_path,
                )
            )
        if not report_instance.data["is_date_skipped"]:
            feedback_objs = feedback_objs.filter(
                created_at__date__gte=report_instance.start_date, created_at__date__lte=report_instance.end_date
            )
        if report_instance.data["user"] or report_instance.data["user_group"]:
            feedback_objs = feedback_objs.filter(
                Q(user_id__in=report_instance.data["user"])
                | Q(user__related_user_groups__id__in=report_instance.data["user_group"])
            )

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self.setup_excel_header(sheet)
            self.populate_excel_sheet(sheet, feedback_objs)
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            uploaded_file = default_storage.save(f"files/{db_name}/report/{report_instance.name}.xlsx", excel_file)
            report_instance.file_url = default_storage.url(uploaded_file)
            report_instance.call_report_emailtask(file_path=uploaded_file, db_name=db_name)
            report_instance.status = BaseUploadStatusChoices.completed
            report_instance.save()
        except Exception:
            report_instance.status = BaseUploadStatusChoices.failed
            report_instance.save()
        return True


class LeaderboardReportGenerationTask(BaseAppTask):
    """Celery task to Generate Leaderboard Report"""

    @staticmethod
    def setup_excel_header(sheet):
        """Function to Populate Headers"""

        return sheet.append(["UserName", "First Name", "Last Name", "Milestone", "Points"])

    @staticmethod
    def populate_excel_sheet(sheet, users):
        """Function to Populate report data"""

        for user in users:
            sheet.append([user.username, user.first_name, user.last_name, "-", user.total_points or "-"])
            for leaderboard in user.related_leaderboard_activities.all():
                sheet.append(
                    [
                        user.username,
                        user.first_name,
                        user.last_name,
                        leaderboard.milestone.name,
                        leaderboard.points or "-",
                    ]
                )

    def run(self, db_name, report_instance_id):
        """Run Handler"""

        from apps.access.models import User
        from apps.my_learning.models import Report

        self.switch_db(db_name)
        self.logger.info(f"Executing ReportGenerationTask on {db_name} database.")

        report_instance = Report.objects.get(id=report_instance_id)

        users = User.objects.filter(related_leaderboard_activities__isnull=False).distinct()

        match report_instance.data["leaderboard_time_param"]:
            case TimePeriodChoices.day:
                users = (
                    users.annotate(date=TruncDate("related_leaderboard_activities__created_at"))
                    .annotate(total_points=Sum("related_leaderboard_activities__points"))
                    .filter(date=datetime.strptime(report_instance.data["time_param"], "%Y-%m-%d"))
                    .order_by("-total_points")
                )
            case TimePeriodChoices.month:
                users = (
                    users.annotate(month=ExtractMonth("related_leaderboard_activities__created_at"))
                    .annotate(total_points=Sum("related_leaderboard_activities__points"))
                    .filter(month=report_instance.data["time_param"])
                    .order_by("-total_points")
                )
            case TimePeriodChoices.year:
                users = (
                    users.annotate(year=ExtractYear("related_leaderboard_activities__created_at"))
                    .annotate(total_points=Sum("related_leaderboard_activities__points"))
                    .filter(year=report_instance.data["time_param"])
                    .order_by("-total_points")
                )
        if not report_instance.data["leaderboard_time_param"]:
            users = users.annotate(total_points=Sum("related_leaderboard_activities__points")).order_by(
                "-total_points"
            )

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self.setup_excel_header(sheet)
            self.populate_excel_sheet(sheet, users)
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            uploaded_file = default_storage.save(f"files/{db_name}/report/{report_instance.name}.xlsx", excel_file)
            report_instance.file_url = default_storage.url(uploaded_file)
            report_instance.call_report_emailtask(file_path=uploaded_file, db_name=db_name)
            report_instance.status = BaseUploadStatusChoices.completed
            report_instance.save()
        except Exception:
            report_instance.status = BaseUploadStatusChoices.failed
            report_instance.save()
        return True
