import os
from datetime import datetime

import openpyxl

from apps.common.tasks import BaseAppTask


class EnrollmentBulkUploadTask(BaseAppTask):
    """Task to Bulk Enroll Users"""

    request_headers = None

    def read_excel_file(self, file_path):
        """function to read excel file"""

        sheet = openpyxl.load_workbook(file_path).active
        # Creating a list header with Column values of active sheet [1st row]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        # Creating a List with dictionary values [{ header : row }]
        list_of_enrollments = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        return list_of_enrollments

    def run(self, file_path, db_name, authenticated_user, **kwargs):
        """Run handler."""

        from django.utils import timezone

        from apps.access.models import User
        from apps.learning.helpers import LEARNING_INSTANCE_MODELS
        from apps.my_learning.config import (
            ActionChoices,
            ApprovalTypeChoices,
            BaseLearningTypeChoices,
            LearningStatusChoices,
        )
        from apps.my_learning.helpers import get_tracker_instance
        from apps.my_learning.models import Enrollment
        from apps.my_learning.tasks import CalendarActivityCreationTask, UserEnrollmentEmailTask
        from apps.tenant_service.middlewares import get_current_db_name

        BASE_LEARNING_TYPES = [
            BaseLearningTypeChoices.course,
            BaseLearningTypeChoices.learning_path,
            BaseLearningTypeChoices.advanced_learning_path,
        ]

        self.switch_db(db_name)
        self.logger.info("Executing EnrollmentBulkUploadTask.")
        list_of_enrollments = self.read_excel_file(file_path=file_path)

        self.request_headers = kwargs.get("request", None)

        authenticated_user = User.objects.filter(pk=authenticated_user).first()
        for enrollments_data in list_of_enrollments:
            if (
                not enrollments_data["UserEmail"]
                or not enrollments_data["Type"]
                or not enrollments_data["Code"]
                or not enrollments_data["EndDate"]
                or not LEARNING_INSTANCE_MODELS.get(enrollments_data["Type"].strip())
            ):
                continue
            try:
                user = User.objects.filter(email=enrollments_data["UserEmail"]).first()
                learning_type = enrollments_data["Type"].strip()
                if "playground" in learning_type:
                    learning_type = learning_type.replace("playground", "assignment")
                learning_instance = (
                    LEARNING_INSTANCE_MODELS[learning_type].objects.filter(code=enrollments_data["Code"]).first()
                )
                if not learning_instance:
                    continue
                if not isinstance(enrollments_data["EndDate"], datetime):
                    enrollments_data["EndDate"] = datetime.strptime(enrollments_data["EndDate"], "%m/%d/%Y")
                data = {
                    "learning_type": learning_type,
                    "start_date": learning_instance.start_date,
                    "end_date": enrollments_data["EndDate"],
                    "action_date": timezone.now().date(),
                    "is_enrolled": True,
                    "approval_type": ApprovalTypeChoices.tenant_admin,
                    "action": ActionChoices.approved,
                    "reason": "Bulk Enrollment Process",
                    "created_by": authenticated_user,
                    "actionee_id": authenticated_user.id if authenticated_user else None,
                }
                instance, created = Enrollment.objects.update_or_create(
                    **{"user": user, learning_type: learning_instance}, defaults=data
                )
                if created:
                    instance.notify_user()
                    instance.call_leaderboard_tasks(is_assigned=True, request_headers=self.request_headers)
                    if instance.learning_type in BASE_LEARNING_TYPES:
                        CalendarActivityCreationTask().run_task(
                            event_type=instance.learning_type,
                            event_instance_id=learning_instance.id,
                            user_ids=user.id,
                            db_name=get_current_db_name(),
                        )
                    UserEnrollmentEmailTask().run_task(
                        user_id=user.id, enrollment_id=instance.id, db_name=get_current_db_name()
                    )
                tracker_instance = get_tracker_instance(user=user, enrollment_instance=instance)
                if tracker_instance:
                    instance.learning_status = LearningStatusChoices.started
                    instance.save()
            except Exception as e:
                self.logger.info(f"Error while processing {enrollments_data['UserEmail']}: {e}")
        try:
            os.remove(file_path)
        except FileNotFoundError:
            pass
        return True


class BulkUnenrollmentTask(BaseAppTask):
    """Task to Unenroll Users"""

    def read_excel_file(self, file_path):
        """function to read excel file"""

        sheet = openpyxl.load_workbook(file_path).active
        # Creating a list header with Column values of active sheet [1st row]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        # Creating a List with dictionary values [{ header : row }]
        list_of_unenrollments = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        return list_of_unenrollments

    def run(self, file_path, db_name, **kwargs):
        """Run Handler"""

        from apps.my_learning.models import Enrollment

        self.switch_db(db_name)
        self.logger.info(f"Executing Bulk Unenrollment task on {db_name}.")
        list_of_unenrollments = self.read_excel_file(file_path)

        for unenroll_data in list_of_unenrollments:
            if not unenroll_data["UserEmail"] or not unenroll_data["Type"] or not unenroll_data["Code"]:
                self.logger.info("Missing required Values : Email, Type or Code")
                continue
            try:
                learning_type = unenroll_data["Type"].strip()
                if "playground" in learning_type:
                    learning_type = learning_type.replace("playground", "assignment")
                filter_params = {
                    "user__email": unenroll_data["UserEmail"],
                    "learning_type": learning_type,
                    f"{learning_type}__code": unenroll_data["Code"],
                }
                enrollment_obj = Enrollment.objects.filter(**filter_params).first()
                if enrollment_obj:
                    enrollment_obj.delete()
                    enrollment_obj.remove_dependencies()
            except Exception as e:
                self.logger.info(f"Error while processing {unenroll_data['UserEmail']}: {e}")
        try:
            os.remove(file_path)
        except FileNotFoundError:
            pass

        return True
