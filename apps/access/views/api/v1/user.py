import json
from io import BytesIO

import openpyxl
from django.db.models import Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone
from django_filters import rest_framework as filters
from openpyxl.styles import Font
from rest_framework import parsers, serializers
from rest_framework.response import Response

from apps.access.config import UserConnectActionChoices
from apps.access.models import User, UserConnection, UserFriendRequest, UserProfilePicture
from apps.access.serializers.v1 import (
    CommonLearningDashboardSerializer,
    SimpleUserReadOnlyModelSerializer,
    UserCreateModelSerializer,
    UserProfilePictureRetrieveSerializer,
    UserReadOnlyModelSerializer,
)
from apps.access.tasks import UserBulkUploadTask
from apps.access_control.config import RoleTypeChoices
from apps.access_control.fixtures import PolicyChoices
from apps.access_control.models import UserGroup
from apps.common.helpers import unpack_dj_choices
from apps.common.serializers import AppModelSerializer, AppReadOnlyModelSerializer, AppSerializer
from apps.common.views.api import AppAPIView, AppCreateAPIView, AppModelCreateAPIViewSet, AppModelListAPIViewSet
from apps.leaderboard.config import MilestoneChoices
from apps.leaderboard.tasks import CommonLeaderboardTask
from apps.learning.config import SubModuleTypeChoices
from apps.learning.models import Assignment, AssignmentGroup, CatalogueRelation, Course, Expert, LearningPath
from apps.my_learning.config import EnrollmentTypeChoices, LearningStatusChoices
from apps.my_learning.models import (
    CAYakshaResult,
    CourseAssessmentTracker,
    CourseSubModuleTracker,
    Enrollment,
    LPAYakshaResult,
    UserLearningPathTracker,
)
from apps.my_learning.serializers.v1 import UserEnrollmentListModelSerializer
from apps.tenant_service.middlewares import get_current_db_name, get_current_tenant_details


class UserListAPIView(AppModelListAPIViewSet):
    """View to list down all the `User`."""

    class UserMultipleChoiceFilter(filters.FilterSet):
        """Filter class to support multiple choices."""

        id = filters.AllValuesMultipleFilter(field_name="id")
        roles__role_type = filters.CharFilter(field_name="roles__role_type")
        related_user_groups = filters.ModelMultipleChoiceFilter(
            field_name="related_user_groups", queryset=UserGroup.objects.alive()
        )
        assignment = filters.ModelMultipleChoiceFilter(
            field_name="related_assignments", queryset=Assignment.objects.alive()
        )
        assignment_group = filters.ModelMultipleChoiceFilter(
            field_name="related_assignments__related_assignment_relations__assignment_group",
            queryset=AssignmentGroup.objects.alive(),
        )

        class Meta:
            fields = [
                "id",
                "roles__role_type",
                "related_user_groups",
                "assignment",
                "assignment_group",
            ]

    serializer_class = UserReadOnlyModelSerializer
    get_object_model = User
    queryset = User.objects.alive()
    all_table_columns = {}
    policy_slug = PolicyChoices.user_management
    user_group_filter_key = "related_user_groups__id__in"
    filterset_class = UserMultipleChoiceFilter
    search_fields = ["related_user_details__user_id_number", "first_name", "last_name", "email"]

    def get_queryset(self):
        """Based on user group filter the queryset returned for manager."""

        queryset = super().get_queryset()
        user = self.get_user()
        group_cu = self.request.query_params.get("group_cu") or None
        group_id = self.request.query_params.get("group") or None
        if not user.current_role:
            queryset = queryset.none()
        elif user.current_role.role_type == RoleTypeChoices.manager:
            queryset = queryset.filter(related_user_groups__id__in=user.related_manager_user_groups.all())
        elif user.is_staff and group_cu:
            queryset = queryset.filter(Q(related_user_groups__id=None) | Q(related_user_groups__id=group_id)).exclude(
                id=user.id
            )
        return queryset


class UserCreateAPIViewSet(AppModelCreateAPIViewSet):
    """Super Admin View to handle `User` creation."""

    queryset = User.objects.alive()
    serializer_class = UserCreateModelSerializer
    policy_slug = PolicyChoices.user_management


class UserProfilePictureUploadAPIView(AppCreateAPIView):
    """View to handle the upload."""

    class _Serializer(AppModelSerializer):
        """Serializer for write."""

        class Meta(AppModelSerializer.Meta):
            model = UserProfilePicture
            fields = ["id", "image"]

    parser_classes = [parsers.MultiPartParser]
    serializer_class = _Serializer

    def post(self, request, *args, **kwargs):
        """Overridden to call celery task."""

        response = super().post(request, *args, **kwargs)
        if response.status_code == 201:
            CommonLeaderboardTask().run_task(
                milestone_names=MilestoneChoices.upload_profile_picture,
                user_id=self.get_user().id,
                db_name=get_current_db_name(),
            )
        return response


class UserTCAgreeAPIView(AppAPIView):
    """View to agree the terms and conditions"""

    def post(self, request, *args, **kwargs):
        """Handle on post."""

        user = self.get_authenticated_user()
        if user:
            user.data["is_tc_agreed"] = True
            user.save()
            return self.send_response()
        return self.send_error_response()


class UserDashboardAPIView(AppAPIView):
    """Api view to get data of user dashboard."""

    class _ExpiringEnrollmentSerializer(UserEnrollmentListModelSerializer):
        """Retrieve serializer for expiring courses."""

        expires_in = serializers.SerializerMethodField()

        class Meta(UserEnrollmentListModelSerializer.Meta):
            fields = UserEnrollmentListModelSerializer.Meta.fields + ["expires_in"]

        def get_expires_in(self, obj):
            """Returns the number of days the course is about to expire."""

            return (obj.end_date - timezone.now()).days

    class _LeaderboardSerializer(AppReadOnlyModelSerializer):
        """Serializer class to retrieve the leaderboard activities."""

        total_points = serializers.SerializerMethodField()
        profile_picture = UserProfilePictureRetrieveSerializer()

        def get_total_points(self, user):
            """Return the total points which is already calculated in the view."""

            return getattr(user, "total_points", 0)

        class Meta(AppReadOnlyModelSerializer.Meta):
            model = User
            fields = [
                "id",
                "name",
                "profile_picture",
                "total_points",
            ]

    class CourseSerializer(CommonLearningDashboardSerializer):
        """Serializer class for course list."""

        class Meta(CommonLearningDashboardSerializer.Meta):
            model = Course

    class LPSerializer(CommonLearningDashboardSerializer):
        """Serializer class for LP list."""

        class Meta(CommonLearningDashboardSerializer.Meta):
            model = LearningPath

    def get(self, *args, **kwargs):
        """Returns the data for learner dashboard."""

        user = self.get_user()
        user_groups = user.related_user_groups.all()
        context = self.get_serializer_context()
        catalogue_list = (
            CatalogueRelation.objects.filter(Q(user=user) | Q(user_group__in=user_groups))
            .values_list("catalogue_id", flat=True)
            .distinct()
        )
        user_enrolled_objs = Enrollment.objects.filter(Q(user_group__in=user_groups) | Q(user=user), is_enrolled=True)
        user_course_trackers = user.related_user_course_trackers.all()
        completed_courses = user_course_trackers.filter(is_completed=True)
        completed_lps = UserLearningPathTracker.objects.filter(user=user, is_completed=True)
        total_course_duration = (
            completed_courses.aggregate(total_course_duration=Sum("course__duration"))["total_course_duration"] or 0
        )
        total_lp_duration = (
            completed_lps.aggregate(
                total_lp_duration=Sum("learning_path__related_learning_path_courses__course__duration")
            )["total_lp_duration"]
            or 0
        )
        skill_count = (
            completed_courses.values_list("course__skill").count()
            + completed_lps.values_list("learning_path__skill").count()
        )
        new_courses = self.CourseSerializer(
            Course.objects.active()
            .filter(related_learning_catalogues__id__in=catalogue_list)
            .order_by("-created_at")[:3],
            many=True,
        ).data
        ongoing_courses = UserEnrollmentListModelSerializer(
            user_enrolled_objs.filter(
                ~Q(learning_status=LearningStatusChoices.completed),
                course_id__in=user_course_trackers.filter(is_completed=False).values_list("course_id", flat=True),
            ).order_by("-created_at")[:3],
            context=context,
            many=True,
        ).data
        upcoming_courses = UserEnrollmentListModelSerializer(
            user_enrolled_objs.filter(learning_type=EnrollmentTypeChoices.course, course__is_deleted=False)
            .exclude(course__in=user_course_trackers.values_list("course_id", flat=True))
            .order_by("-created_at")[:3],
            context=context,
            many=True,
        ).data
        expiring_courses = (
            self._ExpiringEnrollmentSerializer(
                user_enrolled_objs.filter(
                    ~Q(learning_status=LearningStatusChoices.completed),
                    learning_type=EnrollmentTypeChoices.course,
                    end_date__gte=timezone.now(),
                ).order_by("end_date")[:3],
                context=context,
                many=True,
            )
        ).data
        new_lps = self.LPSerializer(
            LearningPath.objects.active()
            .filter(related_learning_catalogues__id__in=catalogue_list)
            .order_by("-created_at")[:3],
            many=True,
        ).data
        user_objs = User.objects.annotate(total_points=Coalesce(Sum("related_leaderboard_activities__points"), 0))
        my_points = user_objs.get(id=user.id).total_points
        leaderboard = (
            self._LeaderboardSerializer(
                user_objs.order_by("-total_points").exclude(is_superuser=True)[:8],
                many=True,
            )
        ).data
        sub_module_file_submissions = CourseSubModuleTracker.objects.filter(
            sub_module__type=SubModuleTypeChoices.file_submission, module_tracker__course_tracker__user=user
        )
        completed_assignments = (
            user.related_assignment_trackers.filter(is_completed=True).count()
            + sub_module_file_submissions.filter(is_completed=True).count()
        )
        incomplete_assignments = (
            user.related_assignment_trackers.filter(is_completed=False).count()
            + sub_module_file_submissions.filter(is_completed=False).count()
        )
        ca_results = CAYakshaResult.objects.filter(
            Q(schedule__tracker__course_tracker__user=user)
            | Q(schedule__tracker__module_tracker__course_tracker__user=user)
        )
        passed_ca_schedules = (
            ca_results.filter(is_pass=True)
            .order_by("schedule_id")
            .distinct("schedule_id")
            .values_list("schedule_id", flat=True)
        )
        passed_ca_results = passed_ca_schedules.count()
        failed_ca_results = (
            ca_results.filter(is_pass=False)
            .order_by("schedule_id")
            .distinct("schedule_id")
            .exclude(schedule_id__in=passed_ca_schedules)
            .count()
        )
        lpa_results = LPAYakshaResult.objects.filter(schedule__tracker__user=user)
        passed_lpa_schedules = (
            lpa_results.filter(is_pass=True)
            .order_by("schedule_id")
            .distinct("schedule_id")
            .values_list("schedule_id", flat=True)
        )
        passed_lpa_results = passed_lpa_schedules.count()
        failed_lpa_results = (
            lpa_results.filter(is_pass=False)
            .order_by("schedule_id")
            .distinct("schedule_id")
            .exclude(schedule_id__in=passed_lpa_schedules)
            .count()
        )
        resubmit_count = 0
        ca_trackers = CourseAssessmentTracker.objects.filter(
            Q(module_tracker__course_tracker__user=user) | Q(course_tracker__user=user)
        ).exclude(reattempt_count=0)
        for ca_tracker in ca_trackers:
            resubmit_count += ca_tracker.reattempt_count - ca_tracker.available_attempt
        lpa_trackers = user.related_lp_assessment_trackers.exclude(reattempt_count=0)
        for lpa_tracker in lpa_trackers:
            resubmit_count += lpa_tracker.reattempt_count - lpa_tracker.available_attempt
        # TODO: Its hardcoded for demotechademy tenant as per @dhakshan's context. Need to remove this.
        tenant_details = get_current_tenant_details()
        if tenant_details["idp_id"] == 9:
            return self.send_response(
                data={
                    "course": {
                        "new_courses": new_courses,
                        "ongoing_courses": ongoing_courses,
                        "upcoming_courses": upcoming_courses,
                        "expiring_courses": expiring_courses,
                    },
                    "learning_path": {
                        "new_lps": new_lps,
                    },
                    "achievements": {
                        "course_count": 4,
                        "lp_count": 2,
                        "completed_duration": 23 * 3600,
                        "skill_count": 5,
                    },
                    "leaderboard": {
                        "organization_leaderboard": leaderboard,
                    },
                    "assessments": {
                        "passed_count": 3,
                        "failed_count": 0,
                        "resubmit_count": 1,
                    },
                    "assignments": {
                        "completed_count": 1,
                        "incomplete_count": 2,
                    },
                }
            )

        return self.send_response(
            data={
                "course": {
                    "new_courses": new_courses,
                    "ongoing_courses": ongoing_courses,
                    "upcoming_courses": upcoming_courses,
                    "expiring_courses": expiring_courses,
                },
                "learning_path": {
                    "new_lps": new_lps,
                },
                "achievements": {
                    "course_count": completed_courses.count(),
                    "lp_count": completed_lps.count(),
                    "completed_duration": total_course_duration + total_lp_duration,
                    "skill_count": skill_count,
                },
                "leaderboard": {
                    "my_points": my_points,
                    "organization_leaderboard": leaderboard,
                },
                "assessments": {
                    "passed_count": passed_ca_results + passed_lpa_results,
                    "failed_count": failed_ca_results + failed_lpa_results,
                    "resubmit_count": resubmit_count,
                },
                "assignments": {
                    "completed_count": completed_assignments,
                    "incomplete_count": incomplete_assignments,
                },
            }
        )


class UserConnectAPIView(AppAPIView):
    """View to perform user connect actions."""

    class _Serializer(AppSerializer):
        """Serializer class for the same view."""

        user = serializers.PrimaryKeyRelatedField(queryset=User.objects.alive(), required=True)
        action = serializers.ChoiceField(choices=UserConnectActionChoices.choices)

        def validate_user(self, user):
            """Overriden to validate the target user is not the requested user."""

            if user == self.get_user():
                raise serializers.ValidationError({"user": "You are not allowed to connect with yourself."})
            return user

    serializer_class = _Serializer

    def get(self, request, *args, **kwargs):
        """Return meta data."""

        return self.send_response(data=unpack_dj_choices(UserConnectActionChoices.choices))

    def post(self, request, *args, **kwargs):
        """Handle on post."""

        validated_data = self.get_valid_serializer().validated_data
        action = validated_data["action"]
        user_obj = validated_data["user"]
        user = self.get_user()
        target_instance = UserConnection.objects.get_or_create(user=user_obj)[0]
        if action == UserConnectActionChoices.follow:
            target_instance.followers.add(user)
            response_data = {"message": "Successfully followed the user."}
        elif action == UserConnectActionChoices.unfollow:
            target_instance.followers.remove(user)
            response_data = {"message": "Successfully unfollowed the user."}
        elif action == UserConnectActionChoices.accept:
            friend_req = UserFriendRequest.objects.get(from_user=user_obj, to_user=user)
            friend_req.delete()
            requester_instance = UserConnection.objects.get_or_create(user=user)[0]
            requester_instance.friends.add(user_obj)
            target_instance.friends.add(user)
            response_data = {"message": "Friend request accepted successfully."}
        elif action == UserConnectActionChoices.reject:
            if friend_req := UserFriendRequest.objects.filter(from_user=user, to_user=user_obj).first():
                friend_req.delete()
                response_data = {"message": "Friend request rejected successfully."}
            elif friend_req := UserFriendRequest.objects.filter(from_user=user_obj, to_user=user).first():
                friend_req.delete()
                response_data = {"message": "Friend request rejected successfully."}
        elif action == UserConnectActionChoices.add_friend:
            friend_req, created = UserFriendRequest.objects.get_or_create(from_user=user, to_user=user_obj)
            if created:
                response_data = {"message": "Friend request sent successfully."}
            else:
                response_data = {"message": "Friend request already sent."}

        return self.send_response(data=response_data)


class UserConnectionListAPIViewSet(AppModelListAPIViewSet):
    """View to list down all the `User`."""

    class _Serializer(SimpleUserReadOnlyModelSerializer):
        """Serializer class for the same view"""

        is_friend = serializers.SerializerMethodField()
        is_following = serializers.SerializerMethodField()
        request_type = serializers.SerializerMethodField()
        is_expert = serializers.SerializerMethodField()

        def get_is_friend(self, obj):
            """Returns True if the current user is friend of the requested object."""

            if obj := UserConnection.objects.filter(user=obj).first():
                return bool(obj.friends.filter(id=self.get_user().id))
            return False

        def get_is_following(self, obj):
            """Returns True if the current user is following the requested object."""

            if obj := UserConnection.objects.filter(user=obj).first():
                return bool(obj.followers.filter(id=self.get_user().id))
            return False

        def get_request_type(self, obj):
            """Returns the request action type."""

            user = self.get_user()
            if UserFriendRequest.objects.filter(from_user=user, to_user=obj):
                return "sent"
            elif UserFriendRequest.objects.filter(from_user=obj, to_user=user):
                return "recieved"
            return None

        def get_is_expert(self, obj):
            """Returns True if the current user is expert or not."""

            return bool(Expert.objects.filter(user=obj))

        class Meta(SimpleUserReadOnlyModelSerializer.Meta):
            fields = SimpleUserReadOnlyModelSerializer.Meta.fields + [
                "is_friend",
                "is_following",
                "request_type",
                "is_expert",
            ]

    serializer_class = _Serializer
    get_object_model = User
    queryset = User.objects.alive().order_by("created_at")
    search_fields = ["first_name", "last_name"]

    def get_queryset(self):
        """Overriden to filter the queryset based on the parameters passed."""

        user = self.get_user()
        queryset = super().get_queryset().exclude(id=user.id)
        instance = UserConnection.objects.get_or_create(user=user)[0]
        if query_params := self.request.query_params.get("filter"):
            if query_params == "all":
                return queryset
            elif query_params == "followers":
                return instance.followers.all()
            elif query_params == "following":
                return queryset.filter(related_user_connections__followers=user)
            elif query_params == "friends":
                return instance.friends.all()
            elif query_params == "friend_request":
                if request_type := self.request.query_params.get("request_type"):
                    if request_type == "sent":
                        return queryset.filter(related_respond_users__from_user=user)
                    elif request_type == "recieved":
                        return queryset.filter(related_request_users__to_user=user)
        return queryset.none()


class UserBulkUploadAPIView(AppAPIView):
    """API View to upload Bulk users from xlsx file"""

    def post(self, request, *args, **kwargs):
        """Handle on post"""

        if not (uploaded_file := request.FILES.get("file")):
            return self.send_error_response(data="File not found")
        if not uploaded_file.name.endswith(".xlsx"):
            return self.send_error_response(data="Unsupported file format. Please upload a .xlsx file.")
        # Opening active sheet of the uploaded file
        sheet = openpyxl.load_workbook(uploaded_file).active
        # Creating a list header with Column values of active sheet [1st row]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        # Creating a List with dictionary values [{ header : row }]
        list_of_users = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        UserBulkUploadTask().run_task(list_of_users=list_of_users, db_name=get_current_db_name())
        return self.send_response(data={"message": "User Bulk Upload Successfull"})


class UserBulkUploadSampleFileAPIView(AppAPIView):
    """API to generate a Sample xlsx file for User Bulk Upload"""

    def get(self, request, *args, **kwargs):
        """Handle on GET"""

        buffer = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "User_Bulk_Upload"
        data = (
            [
                "Tenant Display Name",
                "Tenant Name",
                "Employee Id",
                "Role",
                "Email",
                "Skill",
                "First Name",
                "Last Name",
                "Employment Status",
                "Start Date",
                "Active",
                "Job Description",
                "Job Title",
                "Department Code",
                "Department Title",
                "Current Manager Name",
                "Current Manager Employee Id",
                "Current Manager Email",
                "Manager 2 Email",
                "Manager 3 Email",
                "Config Json",
                "Business Unit/User Group",
                "User Id Number",
                "User Grade",
                "Is Onsite User",
                "City",
                "State",
                "Country",
            ],
            [
                "testpro",
                "testpro",
                "ab12cd45",
                "Learner",
                "harpreetds@gmail.com",
                "Selenium Automation",
                "renovuser",
                "C",
                "Employment - Full Time",
                "mm/dd/yyyy",
                1,
                "L1",
                "Associate Lead",
                "SSG",
                "Admin & Facility Management",
                "Ramaro",
                95,
                "ramarao.bolli@prolifics.com",
                "ramarao.bolli@prolifics.com",
                "Subbaraju.Patchamatla@prolifics.com",
                json.dumps({"config": "json"}),
                "Business Unit",
                "67ef89gh",
                "C",
                1,
                "Chennai",
                "Tamil Nadu",
                "India",
            ],
        )
        for row in data:
            worksheet.append(row)
        cell_to_make_bold = {
            "B1": "Tenant Name",
            "D1": "Role",
            "E1": "Email",
            "G1": "First Name",
        }
        for cell, value in cell_to_make_bold.items():
            worksheet[cell] = value
            worksheet[cell].font = Font(bold=True)
        workbook.save(buffer)
        buffer.seek(0)
        response = Response(
            headers={"Content-Disposition": "attachment; filename=User_Bulk_Upload_Template.xlsx"},
            content_type="application/ms-excel",
        )
        response.content = buffer.read()
        return response
